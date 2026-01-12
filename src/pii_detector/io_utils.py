"""Input/output helpers for CSV and XLSX."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_COLUMNS = ["ID", "Texto Mascarado"]


@dataclass
class CsvData:
    rows: List[Dict[str, str]]
    fieldnames: List[str]
    delimiter: str


@dataclass
class XlsxData:
    workbook: object
    worksheet: Worksheet
    rows: List[Dict[str, str]]
    header: List[str]
    header_row: int


class InputError(Exception):
    pass


def _validate_columns(columns: List[str]) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in columns]
    if missing:
        raise InputError(f"Colunas obrigatórias ausentes: {', '.join(missing)}")


def read_csv(path: Path) -> CsvData:
    text = path.read_text(encoding="utf-8")
    sample = text[:2048]
    try:
        delimiter = csv.Sniffer().sniff(sample).delimiter
    except csv.Error:
        delimiter = ","

    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            raise InputError("Arquivo CSV sem cabeçalho válido")
        fieldnames = list(reader.fieldnames)
        _validate_columns(fieldnames)
        for row in reader:
            rows.append({key: str(value) if value is not None else "" for key, value in row.items()})

    return CsvData(rows=rows, fieldnames=fieldnames, delimiter=delimiter)


def _find_target_sheet(workbook) -> Tuple[Worksheet, int, List[str]]:
    for sheet in workbook.worksheets:
        header_row = 1
        header_cells = [cell.value for cell in sheet[header_row]]
        header = [str(value).strip() if value is not None else "" for value in header_cells]
        if all(col in header for col in REQUIRED_COLUMNS):
            return sheet, header_row, header
    raise InputError("Nenhuma aba contém as colunas obrigatórias 'ID' e 'Texto Mascarado'")


def read_xlsx(path: Path) -> XlsxData:
    workbook = load_workbook(path)
    worksheet, header_row, header = _find_target_sheet(workbook)
    _validate_columns(header)

    rows: List[Dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        row_dict: Dict[str, str] = {}
        for idx, column in enumerate(header):
            if not column:
                continue
            value = row[idx] if idx < len(row) else ""
            row_dict[column] = str(value) if value is not None else ""
        rows.append(row_dict)

    return XlsxData(
        workbook=workbook,
        worksheet=worksheet,
        rows=rows,
        header=header,
        header_row=header_row,
    )


def write_csv(path: Path, data: CsvData, output_rows: List[Dict[str, str]], output_columns: List[str]) -> None:
    fieldnames = list(data.fieldnames)
    for col in output_columns:
        if col not in fieldnames:
            fieldnames.append(col)

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=data.delimiter)
        writer.writeheader()
        for row in output_rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_xlsx(path: Path, data: XlsxData, output_rows: List[Dict[str, str]], output_columns: List[str]) -> None:
    header = list(data.header)
    for col in output_columns:
        if col not in header:
            header.append(col)

    sheet = data.worksheet
    for idx, col in enumerate(header, start=1):
        sheet.cell(row=data.header_row, column=idx, value=col)

    for row_idx, row in enumerate(output_rows, start=data.header_row + 1):
        for col_idx, col in enumerate(header, start=1):
            value = row.get(col, "")
            sheet.cell(row=row_idx, column=col_idx, value=value)

    data.workbook.save(path)
